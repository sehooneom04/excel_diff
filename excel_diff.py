"""
excel_diff.py — 수정본을 그대로 복사한 뒤 변경된 셀에만 색상 표시

사용법:
    python excel_diff.py 원본.xlsx 수정본.xlsx
    python excel_diff.py 원본.xlsx 수정본.xlsx -o 결과.xlsx

설치:
    pip install openpyxl
"""

import argparse
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.comments import Comment


FILL_ADDED     = PatternFill("solid", fgColor="CCFFCC")  # 값 생김 (연초록)
FILL_REMOVED   = PatternFill("solid", fgColor="FFCCCC")  # 값 사라짐 (연빨강)
FILL_CHANGED   = PatternFill("solid", fgColor="FFD966")  # 값 변경 (진노랑)
FILL_ROW_CHG   = PatternFill("solid", fgColor="FFF9E6")  # 변경 행 나머지 (연노랑)
FILL_ROW_ADDED = PatternFill("solid", fgColor="E6FFE6")  # 추가된 행 전체 (연초록)


def read_sheet(ws) -> list:
    if ws is None:
        return []
    rows = []
    for row in ws.iter_rows():
        rows.append([
            "" if cell.value is None else str(cell.value).strip()
            for cell in row
        ])
    return rows


def get_val(data, r, c):
    if r >= len(data) or c >= len(data[r]):
        return ""
    return data[r][c]


def get_master_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for merge_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return ws.cell(row=min_row, column=min_col)
    return None


def safe_fill(ws, row, col, fill):
    cell = get_master_cell(ws, row, col)
    if cell is not None and not isinstance(cell, MergedCell):
        cell.fill = fill


def safe_fill_comment(ws, row, col, fill, text):
    cell = get_master_cell(ws, row, col)
    if cell is not None and not isinstance(cell, MergedCell):
        cell.fill = fill
        cell.comment = Comment(text, "diff")


def diff_excel(file1: Path, file2: Path, output: Path):
    shutil.copy2(file2, output)

    wb1    = openpyxl.load_workbook(file1, data_only=False)
    wb2    = openpyxl.load_workbook(file2, data_only=False)
    wb_out = openpyxl.load_workbook(output, data_only=False)

    all_stats = {}

    for sheet_name in wb2.sheetnames:
        ws1    = wb1[sheet_name] if sheet_name in wb1.sheetnames else None
        ws_out = wb_out[sheet_name]

        data1 = read_sheet(ws1)
        data2 = read_sheet(wb2[sheet_name])

        max_rows = max(len(data1), len(data2))
        max_cols = max(
            max((len(r) for r in data1), default=0),
            max((len(r) for r in data2), default=0),
        )

        stats = {"added": 0, "removed": 0, "changed": 0}

        for r in range(max_rows):
            in1     = r < len(data1)
            in2     = r < len(data2)
            excel_r = r + 1

            # ── 수정본에만 있는 행 전체: 추가 ─────────────────────────
            if in2 and not in1:
                for c in range(max_cols):
                    safe_fill(ws_out, excel_r, c + 1, FILL_ROW_ADDED)
                stats["added"] += 1

            # ── 원본에만 있는 행: 삭제 (수정본에 없으므로 카운트만)
            elif in1 and not in2:
                stats["removed"] += 1

            # ── 양쪽 모두 있는 행: 셀 단위 비교 ──────────────────────
            else:
                changed_cols = [
                    c for c in range(max_cols)
                    if get_val(data1, r, c) != get_val(data2, r, c)
                ]
                if changed_cols:
                    row_has_change = False
                    for c in changed_cols:
                        old = get_val(data1, r, c)
                        new = get_val(data2, r, c)

                        if old == "" and new != "":
                            # 공란 → 값: 초록
                            safe_fill_comment(
                                ws_out, excel_r, c + 1,
                                FILL_ADDED, f"추가됨: {new}"
                            )
                        elif old != "" and new == "":
                            # 값 → 공란: 빨강
                            safe_fill_comment(
                                ws_out, excel_r, c + 1,
                                FILL_REMOVED, f"삭제됨: {old}"
                            )
                        else:
                            # 값 → 다른 값: 노랑
                            safe_fill_comment(
                                ws_out, excel_r, c + 1,
                                FILL_CHANGED, f"변경 전: {old}"
                            )
                        row_has_change = True

                    if row_has_change:
                        # 변경 없는 나머지 셀에 연노랑 배경
                        unchanged = [c for c in range(max_cols) if c not in changed_cols]
                        for c in unchanged:
                            safe_fill(ws_out, excel_r, c + 1, FILL_ROW_CHG)
                        stats["changed"] += 1

        all_stats[sheet_name] = stats
        print(f"  [{sheet_name}]  +{stats['added']} 추가  -{stats['removed']} 삭제  ~{stats['changed']} 변경")

    # ── 요약 시트 ─────────────────────────────────────────────────────────
    ws_sum = wb_out.create_sheet("_요약", 0)
    hfill  = PatternFill("solid", fgColor="2D3748")
    hfont  = Font(color="FFFFFF", bold=True)

    headers = ["시트명", "추가된 행", "삭제된 행", "변경된 행", "합계"]
    ws_sum.append(headers)
    for c in range(1, 6):
        cell = ws_sum.cell(row=1, column=c)
        cell.fill = hfill
        cell.font = hfont
        ws_sum.column_dimensions[get_column_letter(c)].width = 20

    for sname, s in all_stats.items():
        ws_sum.append([sname, s["added"], s["removed"], s["changed"],
                       s["added"] + s["removed"] + s["changed"]])

    ws_sum.append([])
    ws_sum.append(["색상", "설명"])
    legend = [
        ("연초록 셀", "공란 → 값 (새로 생긴 셀)",       FILL_ADDED),
        ("연빨강 셀", "값 → 공란 (사라진 셀)",           FILL_REMOVED),
        ("진노랑 셀", "값 → 다른 값 (변경된 셀)",        FILL_CHANGED),
        ("연노랑 행", "변경이 있는 행의 나머지 셀",       FILL_ROW_CHG),
        ("연초록 행", "수정본에서 새로 추가된 행 전체",   FILL_ROW_ADDED),
    ]
    for label, desc, fill in legend:
        ws_sum.append([label, desc])
        ws_sum.cell(row=ws_sum.max_row, column=1).fill = fill

    wb_out.save(output)
    total = sum(s["added"] + s["removed"] + s["changed"] for s in all_stats.values())
    print(f"\n✅ 완료: {output}  (총 {total}개 변경)")


def main():
    parser = argparse.ArgumentParser(
        description="두 엑셀 파일 비교 — 수정본 서식 그대로 유지하며 변경 셀만 표시"
    )
    parser.add_argument("file1", help="원본 파일 (.xlsx)")
    parser.add_argument("file2", help="수정본 파일 (.xlsx)")
    parser.add_argument("-o", "--output", default="diff_result.xlsx",
                        help="출력 파일명 (기본값: diff_result.xlsx)")
    args = parser.parse_args()

    f1, f2, out = Path(args.file1), Path(args.file2), Path(args.output)
    if not f1.exists(): print(f"오류: {f1} 없음"); sys.exit(1)
    if not f2.exists(): print(f"오류: {f2} 없음"); sys.exit(1)

    print(f"비교 중: {f1.name}  vs  {f2.name}")
    diff_excel(f1, f2, out)


if __name__ == "__main__":
    main()