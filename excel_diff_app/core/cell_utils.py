# filepath: excel_diff_app/core/cell_utils.py
"""셀 관련 유틸리티 함수"""

from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.comments import Comment


def get_master_cell(ws, row, col):
    """병합된 셀의 경우 마스터 셀 반환"""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for merge_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return ws.cell(row=min_row, column=min_col)
    return None


def safe_fill(ws, row, col, fill):
    """셀에 배경색 적용 (병합 셀 처리)"""
    cell = get_master_cell(ws, row, col)
    if cell is not None and not isinstance(cell, MergedCell):
        cell.fill = fill


def safe_fill_comment(ws, row, col, fill, text):
    """셀에 배경색과 코멘트 적용 (병합 셀 처리)"""
    cell = get_master_cell(ws, row, col)
    if cell is not None and not isinstance(cell, MergedCell):
        cell.fill = fill
        cell.comment = Comment(text, "diff")


def get_val(data, r, c):
    """2D 데이터에서 값 가져오기 (안전하게)"""
    if r >= len(data) or c >= len(data[r]):
        return ""
    return data[r][c]


def get_column_letter_by_index(col_idx):
    """0-based 인덱스를 Excel 열 문자로 변환"""
    return get_column_letter(col_idx + 1)