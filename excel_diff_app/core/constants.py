# filepath: excel_diff_app/core/constants.py
"""상수 및 설정값"""

from openpyxl.styles import PatternFill

# 색상 정의
FILL_ADDED     = PatternFill("solid", fgColor="CCFFCC")  # 값 생김 (연초록)
FILL_REMOVED   = PatternFill("solid", fgColor="FFCCCC")  # 값 사라짐 (연빨강)
FILL_CHANGED   = PatternFill("solid", fgColor="FFD966")  # 값 변경 (진노랑)
FILL_ROW_CHG   = PatternFill("solid", fgColor="FFF9E6")  # 변경 행 나머지 (연노랑)
FILL_ROW_ADDED = PatternFill("solid", fgColor="E6FFE6")  # 추가된 행 전체 (연초록)

# 헤더 색상
HEADER_FILL = PatternFill("solid", fgColor="2D3748")
HEADER_FONT_COLOR = "FFFFFF"

# 기본 출력 파일명
DEFAULT_OUTPUT = "diff_result.xlsx"

# 파일 필터
FILE_FILTERS = [
    ("Excel 파일", "*.xlsx *.xls"),
    ("모든 파일", "*.*")
]