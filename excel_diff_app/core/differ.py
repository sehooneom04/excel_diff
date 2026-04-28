# filepath: excel_diff_app/core/differ.py
"""Excel 비교 핵심 로직"""

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from .constants import (
    FILL_ADDED, FILL_REMOVED, FILL_CHANGED, FILL_ROW_CHG, FILL_ROW_ADDED,
    HEADER_FILL, HEADER_FONT_COLOR
)
from .reader import read_sheet, load_workbook_safe, get_sheet
from .cell_utils import safe_fill, safe_fill_comment, get_val


def diff_excel(file1: Path, file2: Path, output: Path, progress_callback=None):
    """
    두 Excel 파일을 비교하여 변경 사항을 표시한 결과 파일 생성
    
    Args:
        file1: 원본 파일 경로
        file2: 수정본 파일 경로
        output: 출력 파일 경로
        progress_callback: 진행률 콜백 함수 (optional)
    
    Returns:
        dict: 시트별 통계 정보
    """
    shutil.copy2(file2, output)

    wb1    = load_workbook_safe(file1, data_only=False)
    wb2    = load_workbook_safe(file2, data_only=False)
    wb_out = load_workbook_safe(output, data_only=False)

    all_stats = {}

    sheet_names = wb2.sheetnames
    total_sheets = len(sheet_names)

    for idx, sheet_name in enumerate(sheet_names):
        if progress_callback:
            progress_callback(idx, total_sheets, sheet_name)
        
        ws1    = get_sheet(wb1, sheet_name)
        ws_out = wb_out[sheet_name]

        data1 = read_sheet(ws1)
        data2 = read_sheet(wb2[sheet_name])

        max_rows = max(len(data1), len(data2))
        max_cols = max(
            max((len(r) for r in data1), default=0),
            max((len(r) for r in data2), default=0),
        )

        stats = {"added": 0, "removed": 0, "changed": 0}

        for r in range(max_rows):
            in1     = r < len(data1)
            in2     = r < len(data2)
            excel_r = r + 1

            # 수정본에만 있는 행 전체: 추가
            if in2 and not in1:
                for c in range(max_cols):
                    safe_fill(ws_out, excel_r, c + 1, FILL_ROW_ADDED)
                stats["added"] += 1

            # 원본에만 있는 행: 삭제 (수정본에 없으므로 카운트만)
            elif in1 and not in2:
                stats["removed"] += 1

            # 양쪽 모두 있는 행: 셀 단위 비교
            else:
                changed_cols = [
                    c for c in range(max_cols)
                    if get_val(data1, r, c) != get_val(data2, r, c)
                ]
                if changed_cols:
                    row_has_change = False
                    for c in changed_cols:
                        old = get_val(data1, r, c)
                        new = get_val(data2, r, c)

                        if old == "" and new != "":
                            # 공란 → 값: 초록
                            safe_fill_comment(
                                ws_out, excel_r, c + 1,
                                FILL_ADDED, f"추가됨: {new}"
                            )
                        elif old != "" and new == "":
                            # 값 → 공란: 빨강
                            safe_fill_comment(
                                ws_out, excel_r, c + 1,
                                FILL_REMOVED, f"삭제됨: {old}"
                            )
                        else:
                            # 값 → 다른 값: 노랑
                            safe_fill_comment(
                                ws_out, excel_r, c + 1,
                                FILL_CHANGED, f"변경 전: {old}"
                            )
                        row_has_change = True

                    if row_has_change:
                        # 변경 없는 나머지 셀에 연노랑 배경
                        unchanged = [c for c in range(max_cols) if c not in changed_cols]
                        for c in unchanged:
                            safe_fill(ws_out, excel_r, c + 1, FILL_ROW_CHG)
                        stats["changed"] += 1

        all_stats[sheet_name] = stats

    # 요약 시트 생성
    _create_summary_sheet(wb_out, all_stats)

    wb_out.save(output)
    return all_stats


def _create_summary_sheet(wb, stats):
    """요약 시트 생성"""
    ws_sum = wb.create_sheet("_요약", 0)
    hfill  = HEADER_FILL
    hfont  = Font(color=HEADER_FONT_COLOR, bold=True)

    headers = ["시트명", "추가된 행", "삭제된 행", "변경된 행", "합계"]
    ws_sum.append(headers)
    for c in range(1, 6):
        cell = ws_sum.cell(row=1, column=c)
        cell.fill = hfill
        cell.font = hfont
        ws_sum.column_dimensions[get_column_letter(c)].width = 20

    for sname, s in stats.items():
        ws_sum.append([sname, s["added"], s["removed"], s["changed"],
                       s["added"] + s["removed"] + s["changed"]])

    ws_sum.append([])
    ws_sum.append(["색상", "설명"])
    legend = [
        ("연초록 셀", "공란 → 값 (새로 생긴 셀)",       FILL_ADDED),
        ("연빨강 셀", "값 → 공란 (사라진 셀)",           FILL_REMOVED),
        ("진노랑 셀", "값 → 다른 값 (변경된 셀)",        FILL_CHANGED),
        ("연노랑 행", "변경이 있는 행의 나머지 셀",       FILL_ROW_CHG),
        ("연초록 행", "수정본에서 새로 추가된 행 전체",   FILL_ROW_ADDED),
    ]
    for label, desc, fill in legend:
        ws_sum.append([label, desc])
        ws_sum.cell(row=ws_sum.max_row, column=1).fill = fill


def get_total_changes(stats):
    """총 변경 수 계산"""
    return sum(s["added"] + s["removed"] + s["changed"] for s in stats.values())


def format_stats_message(stats):
    """통계 메시지 포맷"""
    messages = []
    for sheet_name, s in stats.items():
        messages.append(
            f"  [{sheet_name}]  +{s['added']} 추가  -{s['removed']} 삭제  ~{s['changed']} 변경"
        )
    return "\n".join(messages)