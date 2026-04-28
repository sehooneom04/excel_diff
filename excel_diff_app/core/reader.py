# filepath: excel_diff_app/core/reader.py
"""Excel 파일 읽기 모듈"""

from openpyxl import load_workbook


def read_sheet(ws):
    """워크시트에서 데이터를 2D 리스트로 읽기"""
    if ws is None:
        return []
    rows = []
    for row in ws.iter_rows():
        rows.append([
            "" if cell.value is None else str(cell.value).strip()
            for cell in row
        ])
    return rows


def load_workbook_safe(file_path, data_only=False):
    """워크북을 안전하게 로드"""
    return load_workbook(file_path, data_only=data_only)


def get_sheet_names(workbook):
    """워크북에서 모든 시트 이름 반환"""
    return workbook.sheetnames


def get_sheet(workbook, sheet_name):
    """시트 이름으로 시트 객체 반환"""
    return workbook[sheet_name] if sheet_name in workbook.sheetnames else None